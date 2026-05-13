"""Reporting: writes an Excel workbook of the latest snapshot + recommendations,
and emails it via SMTP.
"""
from __future__ import annotations

import logging
import smtplib
import ssl
from dataclasses import asdict, is_dataclass
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger("reporter")


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALERT_FILL = PatternFill("solid", fgColor="C00000")
ALERT_FONT = Font(color="FFFFFF", bold=True)


def _autosize(ws):
    for col in ws.columns:
        col = list(col)
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)


def _write_header(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_report_xlsx(
    out_path: str | Path,
    snapshot,        # scraper.Snapshot
    forecast,        # forecasting.DemandForecast
    policy,          # rop_calculator.PolicyRecommendation
    endgame_rec,     # endgame.EndGameRecommendation
    final_rop: float,
    final_q: float,
    use_truck: bool,
    extra_notes: list[str] | None = None,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Group 3 Watchdog Snapshot"
    ws["A1"].font = Font(bold=True, size=14)

    rows = [
        ("Timestamp", getattr(snapshot, "timestamp", "")),
        ("Current Day", getattr(snapshot, "current_day", None)),
        ("Cash", getattr(snapshot, "cash", None)),
        ("Warehouse Inventory", getattr(snapshot, "warehouse_inventory", None)),
        ("Factory WIP", getattr(snapshot, "factory_wip", None)),
        ("Factory Capacity", getattr(snapshot, "factory_capacity", None)),
        ("In Transit", getattr(snapshot, "in_transit", None)),
        ("Current Order Point (game)", getattr(snapshot, "factory_order_point", None)),
        ("Current Batch Size (game)", getattr(snapshot, "factory_batch_size", None)),
        ("", ""),
        ("End-game Phase", endgame_rec.phase.value),
        ("Days Remaining", endgame_rec.days_remaining),
        ("Expected Remaining Demand", round(endgame_rec.expected_remaining_demand, 1)),
        ("Pipeline + Warehouse Inventory", round(endgame_rec.pipeline_plus_warehouse, 1)),
        ("", ""),
        ("RECOMMENDED Order Point", round(final_rop, 1)),
        ("RECOMMENDED Batch Size (Q)", round(final_q, 1)),
        ("RECOMMENDED Shipping", "Truck" if use_truck else "Mail"),
        ("RECOMMENDED Target Capacity", round(policy.target_capacity, 1)),
        ("RECOMMENDED Capacity Add (drum/day)", round(policy.expansion_drums, 1)),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # Highlight key recommendation rows
    for r in range(3 + 14, 3 + 19):
        ws.cell(row=r, column=1).fill = HEADER_FILL
        ws.cell(row=r, column=1).font = HEADER_FONT
    _autosize(ws)

    # --- Notes sheet ---
    ws = wb.create_sheet("Decision Notes")
    ws["A1"] = "Notes from policy + end-game engines"
    ws["A1"].font = Font(bold=True, size=12)
    notes = list(policy.notes) + [endgame_rec.note]
    if extra_notes:
        notes += extra_notes
    for i, n in enumerate(notes, start=3):
        ws.cell(row=i, column=1, value=n).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30
    ws.column_dimensions["A"].width = 110

    # --- ROP calculation sheet ---
    ws = wb.create_sheet("ROP Calculation")
    _write_header(ws, 1, ["Metric", "Value"])
    rop_rows = [
        ("Day", policy.day),
        ("Service level", "see policy config"),
        ("Reorder point (raw)", round(policy.reorder_point, 2)),
        ("Batch size (raw)", round(policy.batch_size, 2)),
        ("Fills a truck (>=200)?", "Yes" if policy.fills_truck else "No"),
        ("Use truck?", "Yes" if policy.use_truck else "No"),
        ("Target capacity", round(policy.target_capacity, 2)),
        ("Capacity to add (drum/day)", round(policy.expansion_drums, 2)),
    ]
    for i, (k, v) in enumerate(rop_rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    _autosize(ws)

    # --- Forecast preview sheet ---
    ws = wb.create_sheet("Forecast (next 60d)")
    _write_header(ws, 1, ["Day", "Forecast Mean Demand", "Forecast Stdev"])
    horizon_days = forecast.days[:60] if len(forecast.days) > 60 else forecast.days
    horizon_mean = forecast.mean[:60] if len(forecast.mean) > 60 else forecast.mean
    horizon_sd = forecast.stdev[:60] if len(forecast.stdev) > 60 else forecast.stdev
    for i, (d, m, s) in enumerate(zip(horizon_days, horizon_mean, horizon_sd), start=2):
        ws.cell(row=i, column=1, value=int(d))
        ws.cell(row=i, column=2, value=float(m))
        ws.cell(row=i, column=3, value=float(s))
    _autosize(ws)

    # --- Seasonality sheet ---
    ws = wb.create_sheet("Seasonal Index")
    _write_header(ws, 1, ["Day-of-Year (1..365)", "Seasonal Index"])
    for i, v in enumerate(forecast.seasonal_index, start=1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=float(v))
    _autosize(ws)

    # --- Series sheets ---
    series_attrs = [
        ("Demand Series", "demand_series"),
        ("Lost Demand Series", "lost_demand_series"),
        ("Inventory Series", "inventory_series"),
        ("Cash Series", "cash_series"),
        ("Shipments Series", "shipments_series"),
        ("WIP Series", "wip_series"),
    ]
    for title, attr in series_attrs:
        data = getattr(snapshot, attr, None) or []
        if not data:
            continue
        ws = wb.create_sheet(title)
        _write_header(ws, 1, ["Day", "Value"])
        for i, (d, v) in enumerate(data, start=2):
            ws.cell(row=i, column=1, value=float(d))
            ws.cell(row=i, column=2, value=float(v))
        _autosize(ws)

    # --- Standings ---
    if getattr(snapshot, "standing", None):
        ws = wb.create_sheet("Team Standing")
        cols = sorted({k for row in snapshot.standing for k in row.keys()})
        _write_header(ws, 1, cols)
        for r, row in enumerate(snapshot.standing, start=2):
            for c, k in enumerate(cols, start=1):
                ws.cell(row=r, column=c, value=row.get(k))
        _autosize(ws)

    wb.save(out_path)
    log.info("Report saved: %s", out_path)
    return out_path


def send_email(
    smtp_cfg: dict,
    subject: str,
    body: str,
    attachment_path: str | Path | None = None,
) -> dict:
    if not smtp_cfg.get("enabled", False):
        return {"sent": False, "reason": "email.enabled=false"}

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_cfg["from_address"]
    msg["To"] = ", ".join(smtp_cfg["recipients"])
    msg.set_content(body)

    if attachment_path:
        attachment_path = Path(attachment_path)
        with open(attachment_path, "rb") as f:
            data = f.read()
        msg.add_attachment(
            data, maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attachment_path.name,
        )

    host = smtp_cfg["smtp_host"]
    port = int(smtp_cfg["smtp_port"])
    use_ssl = smtp_cfg.get("smtp_use_ssl", True)
    user = smtp_cfg["smtp_user"]
    pw = smtp_cfg["smtp_password"]

    if use_ssl:
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL(host, port, context=ctx, timeout=30) as s:
            s.login(user, pw)
            s.send_message(msg)
    else:
        with smtplib.SMTP(host, port, timeout=30) as s:
            s.starttls()
            s.login(user, pw)
            s.send_message(msg)

    log.info("Email sent to %s", msg["To"])
    return {"sent": True, "recipients": smtp_cfg["recipients"]}


def build_email_body(snapshot, final_rop, final_q, use_truck, endgame_rec, policy, extra_notes=None) -> str:
    lines = [
        "Group 3 Watchdog snapshot",
        "=" * 50,
        f"Timestamp:           {getattr(snapshot, 'timestamp', '?')}",
        f"Current day:         {getattr(snapshot, 'current_day', '?')}",
        f"Cash:                {getattr(snapshot, 'cash', '?')}",
        f"Warehouse inventory: {getattr(snapshot, 'warehouse_inventory', '?')}",
        f"Factory WIP:         {getattr(snapshot, 'factory_wip', '?')}",
        f"Factory capacity:    {getattr(snapshot, 'factory_capacity', '?')}",
        f"In transit:          {getattr(snapshot, 'in_transit', '?')}",
        "",
        f"End-game phase:      {endgame_rec.phase.value}",
        f"Days remaining:      {endgame_rec.days_remaining}",
        f"Expected remaining demand: {endgame_rec.expected_remaining_demand:.1f}",
        f"Pipeline + warehouse:      {endgame_rec.pipeline_plus_warehouse:.1f}",
        "",
        "RECOMMENDATIONS",
        "-" * 50,
        f"Order Point:  {final_rop:.0f}",
        f"Batch Size:   {final_q:.0f}",
        f"Shipping:     {'Truck' if use_truck else 'Mail'}",
        f"Capacity:     target {policy.target_capacity:.0f} drum/day "
        f"({'+%.0f' % policy.expansion_drums if policy.expansion_drums > 0 else 'no change'})",
        "",
        "Notes",
        "-" * 50,
    ]
    for n in policy.notes + [endgame_rec.note] + (extra_notes or []):
        lines.append(f"- {n}")
    return "\n".join(lines)
